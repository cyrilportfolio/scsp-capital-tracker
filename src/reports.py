"""Excel workbook and text report handed to the investor relations team."""

from __future__ import annotations

from pathlib import Path

import pandas as pd

SHEET_TITLES = {
    "synthese": "Synthese",
    "controles": "Controles",
    "anomalies": "Anomalies",
    "associes": "Associes",
    "appels": "Appels de capital",
    "avis": "Avis d'appel",
    "avis_objet": "Objet de l'appel",
    "engagements": "Engagements",
    "portefeuille": "Portefeuille",
    "nav": "NAV trimestrielle",
    "comptes": "Comptes associes",
    "etat_de_compte": "Etat de compte",
    "cascade": "Cascade",
    "carry": "Carry par trimestre",
    "performance": "Performance",
    "hurdle": "Preferred return",
}

MONEY_COLUMNS = {
    "engagement", "appele", "non_appele", "montant", "montant_appel_total",
    "montant_total", "appele_avant", "montant_appele", "appele_apres",
    "solde_ouverture", "contributions", "distributions",
    "quote_part_gain_net", "quote_part_commission", "quote_part_frais",
    "carried_interest", "solde_cloture", "resultat_net_attribue",
    "juste_valeur", "cout_acquisition", "produit_cession",
    "plus_value_latente", "plus_value_realisee", "juste_valeur_portefeuille",
    "tresorerie", "nav_avant_carried", "appels", "gain_net_investissements",
    "commission_gestion", "frais_de_fonctionnement", "resultat_net",
    "nav_ouverture", "variation_nav", "ecart_reconciliation", "ecart",
    "aux_associes", "au_cip", "solde_du_bucket", "bucket",
    "contributions_cumulees", "distributions_cumulees", "valeur_residuelle",
    "preferred_return_acquis", "retour_du_capital", "preferred_return_paye",
    "catch_up", "revenant_aux_associes", "capital_appele",
    "distributions_recues", "valeur_totale", "total_comptes_associes",
    "mouvement", "contributions_non_remboursees", "hurdle_capitalise",
    "hurdle_courant", "hurdle_cumule",
}

RATIO_COLUMNS = {"PIC", "DPI", "RVPI", "TVPI", "multiple", "part", "pct_appele"}
PERCENT_COLUMNS = {"tri_net", "part_pct"}


def _style(worksheet, frame: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="BFBFBF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    worksheet.freeze_panes = "A2"
    if len(frame) and len(frame.columns):
        worksheet.auto_filter.ref = worksheet.dimensions

    for index, column in enumerate(frame.columns, start=1):
        if column in MONEY_COLUMNS:
            fmt = "#,##0.00"
        elif column in RATIO_COLUMNS:
            fmt = "0.000"
        elif column in PERCENT_COLUMNS:
            fmt = "0.00%"
        else:
            fmt = None
        for row in range(2, len(frame) + 2):
            cell = worksheet.cell(row=row, column=index)
            if fmt:
                cell.number_format = fmt
            cell.border = Border(bottom=thin)

        longest = frame[column].astype(str).str.len().max()
        longest = 0 if pd.isna(longest) else int(longest)
        width = min(max(len(str(column)), longest) + 3, 58)
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.print_title_rows = "1:1"


def synthesis_frame(context: dict) -> pd.DataFrame:
    rows = [
        ("Fonds", context["fund"]),
        ("Forme juridique", context["legal_form"]),
        ("Devise", context["currency"]),
        ("Date de reporting", context["as_of"]),
        ("Date d'execution", context["run_date"]),
        ("Engagements totaux", context["commitments"]),
        ("Capital appele", context["paid_in"]),
        ("Engagement non appele", context["unfunded"]),
        ("Distributions cumulees", context["distributions"]),
        ("Juste valeur du portefeuille", context["portfolio_value"]),
        ("Tresorerie", context["cash"]),
        ("NAV du fonds", context["nav"]),
        ("Preferred return acquis", context["preferred"]),
        ("Carried interest accru", context["carried"]),
        ("NAV revenant aux associes", context["nav_partners"]),
        ("DPI", context["DPI"]),
        ("RVPI", context["RVPI"]),
        ("TVPI", context["TVPI"]),
        ("TRI net du fonds", context["irr"]),
        ("Anomalies detectees", context["anomalies"]),
        ("dont bloquantes", context["blocking"]),
    ]
    return pd.DataFrame(rows, columns=["Indicateur", "Valeur"])


def write_workbook(path: Path, sheets: dict, context: dict) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    ordered = [("synthese", synthesis_frame(context))]
    ordered += [(key, frame) for key, frame in sheets.items()]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for key, frame in ordered:
            title = SHEET_TITLES.get(key, key)[:31]
            export = frame.copy()
            for column in export.columns:
                if pd.api.types.is_datetime64_any_dtype(export[column]):
                    export[column] = export[column].dt.strftime("%d/%m/%Y")
            export.to_excel(writer, sheet_name=title, index=False)
            _style(writer.sheets[title], export)
    return path


def write_text_report(path: Path, context: dict, waterfall_steps: pd.DataFrame,
                      summary: pd.DataFrame, anomalies: pd.DataFrame,
                      statement: pd.DataFrame) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "RAPPORT TRIMESTRIEL AUX ASSOCIES",
        "=" * 78,
        f"Fonds              : {context['fund']} ({context['legal_form']})",
        f"Date de reporting  : {context['as_of']}",
        f"Execution          : {context['run_date']}",
        "",
        "CHIFFRES CLES",
        "-" * 78,
        f"Engagements                  {context['commitments']:>18,.2f} EUR",
        f"Capital appele               {context['paid_in']:>18,.2f} EUR",
        f"Engagement non appele        {context['unfunded']:>18,.2f} EUR",
        f"Distributions cumulees       {context['distributions']:>18,.2f} EUR",
        f"NAV du fonds                 {context['nav']:>18,.2f} EUR",
        f"Preferred return acquis      {context['preferred']:>18,.2f} EUR",
        f"Carried interest accru       {context['carried']:>18,.2f} EUR",
        f"DPI / RVPI / TVPI            {context['DPI']} / {context['RVPI']} / {context['TVPI']}",
        f"TRI net du fonds             {context['irr']}",
        "",
        "CASCADE DE REPARTITION (whole-of-fund, base liquidation)",
        "-" * 78,
    ]
    for _, step in waterfall_steps.iterrows():
        lines.append(f"{step['etape']}. {step['clause']:<34} "
                     f"associes {step['aux_associes']:>14,.2f}  "
                     f"CIP {step['au_cip']:>12,.2f}")

    lines += ["", "ETAT DES COMPTES ASSOCIES", "-" * 78]
    for _, row in statement.iterrows():
        lines.append(f"{row['code']:<6} {row['nom'][:34]:<34} "
                     f"solde {row['solde_cloture']:>16,.2f} EUR")

    lines += ["", "CONTROLES", "-" * 78]
    for _, row in summary.iterrows():
        lines.append(f"{row['statut']:<11} {row['severite']:<9} "
                     f"{row['anomalies']:>4}  {row['libelle_controle']}")

    if not anomalies.empty:
        lines += ["", "ANOMALIES", "-" * 78]
        for _, row in anomalies.iterrows():
            when = ("" if pd.isna(row["date"])
                    else pd.Timestamp(row["date"]).strftime("%d/%m/%Y"))
            lines.append(f"  [{row['code_controle']}] {row['objet']:<8} "
                         f"{when:<11} {row['message']}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path
